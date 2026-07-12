import os
import zipfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from bot.config import DATA_TEMP_DIR

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(_BASE_DIR, "..", "..", "templates", "ndfl_2025.xlsx")
TEMPLATE_PATH = os.path.abspath(TEMPLATE_PATH)

BASE_SHEETS = ["Титульный лист", "Раздел 1", "Прил-е к Разделу 1", "Раздел 2"]
GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

DRAWING_FILES = [f"xl/drawings/drawing{i}.xml" for i in range(1, 17)]


async def generate_excel(declaration_id: int, data: dict) -> str:
    last_name = data.get("last_name", "Декларация")
    selected = data.get("selected_deductions", {})
    today = datetime.now().strftime("%d%m%y")
    types = []
    if selected.get("medical"): types.append("мед")
    if selected.get("education"): types.append("обуч")
    if selected.get("property"): types.append("имущ")
    if selected.get("investment"): types.append("инв")
    type_str = "-".join(types) if types else "вычет"
    file_name = f"3НДФЛ_{last_name}_{today}_{type_str}.xlsx"
    excel_path = os.path.join(DATA_TEMP_DIR, file_name)

    wb = load_workbook(TEMPLATE_PATH)

    print_sheets = list(BASE_SHEETS)
    if selected.get("education"): print_sheets.append("Прил.5")
    if selected.get("medical"): print_sheets.append("Прил.5 (продолжение)")
    if selected.get("property"): print_sheets.append("Прил.7")
    if selected.get("investment"):
        print_sheets.append("Прил.5")
        print_sheets.append("Расчет к прил.5")

    page_number = 1
    for sheet_name in print_sheets:
        ws = wb[sheet_name]
        ws.sheet_properties.tabColor = "00CC00"
        _write_page_number(ws, str(page_number).zfill(3), is_title=(sheet_name == "Титульный лист"))
        page_number += 1

    for sheet_name in wb.sheetnames:
        if sheet_name not in print_sheets:
            wb[sheet_name].sheet_properties.tabColor = "C0C0C0"

    _fill_title(wb, data, len(print_sheets))
    _fill_section1(wb, data)
    _fill_return_request(wb, data)
    _fill_section2(wb, data)

    if selected.get("education"): _fill_appendix5_education(wb, data)
    if selected.get("medical"): _fill_appendix5_medical(wb, data)
    if selected.get("property"): _fill_appendix7(wb, data)
    if selected.get("investment"):
        _fill_appendix5_investment(wb, data)
        _fill_calc_appendix5(wb, data)

    wb.active = wb["Титульный лист"]
    wb.save(excel_path)

    _restore_drawings(excel_path)
    return excel_path


def _restore_drawings(filepath: str):
    with zipfile.ZipFile(TEMPLATE_PATH, "r") as z_template:
        with zipfile.ZipFile(filepath, "r") as z_generated:
            tmp_path = filepath + ".tmp"
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as z_out:
                for item in z_generated.infolist():
                    if item.filename in DRAWING_FILES:
                        try:
                            data = z_template.read(item.filename)
                            z_out.writestr(item, data)
                        except Exception:
                            z_out.writestr(item, z_generated.read(item.filename))
                    else:
                        z_out.writestr(item, z_generated.read(item.filename))
    os.replace(tmp_path, filepath)


def _safe_write(ws, cell_ref, value, font_size=None):
    value = str(value) if value is not None else ""
    cell = ws[cell_ref]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                parent_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                parent_cell.value = value
                if font_size:
                    parent_cell.font = Font(size=font_size)
                return
    else:
        cell.value = value
        if font_size:
            cell.font = Font(size=font_size)


def _write_fio_field(ws, text, start_col, row):
    text = str(text) if text else ""
    col = start_col
    for char in text:
        if col > 60: break
        col_letter = get_column_letter(col)
        _safe_write(ws, f"{col_letter}{row}", char.upper())
        col += 2


def _write_number_field(ws, text, start_col, row):
    text = str(text) if text else ""
    col = start_col
    for char in text:
        col_letter = get_column_letter(col)
        _safe_write(ws, f"{col_letter}{row}", char)
        col += 1


def _write_amount_with_kopeks(ws, amount, start_col, row):
    amount_str = f"{amount:.2f}"
    parts = amount_str.split(".")
    _write_number_field(ws, parts[0], start_col, row)
    _safe_write(ws, f"AM{row}", parts[1][0])
    _safe_write(ws, f"AN{row}", parts[1][1])


def _write_page_number(ws, number_str, is_title=False):
    number_str = str(number_str).zfill(3)
    if is_title:
        _safe_write(ws, "AU4", number_str[0])
        _safe_write(ws, "AW4", number_str[1])
        _safe_write(ws, "AY4", number_str[2])
    else:
        _safe_write(ws, "X4", number_str[0])
        _safe_write(ws, "Y4", number_str[1])
        _safe_write(ws, "Z4", number_str[2])


def _write_fio_section_header(ws, data):
    last_name = str(data.get("last_name", ""))
    _safe_write(ws, "E7", last_name.upper())
    first_name = str(data.get("first_name", ""))
    if first_name: _safe_write(ws, "AH7", first_name[0].upper())
    middle_name = str(data.get("middle_name", ""))
    if middle_name and middle_name != "-": _safe_write(ws, "AK7", middle_name[0].upper())


# ==================== ТИТУЛЬНЫЙ ЛИСТ ====================

def _fill_title(wb, data, total_pages):
    ws = wb["Титульный лист"]
    _write_inn(ws, data.get("taxpayer_inn", ""))
    _safe_write(ws, "K11", "0"); _safe_write(ws, "M11", "0"); _safe_write(ws, "O11", "0")
    _safe_write(ws, "AC11", "3"); _safe_write(ws, "AE11", "4")
    year = str(data.get("year", ""))
    if len(year) >= 4:
        _safe_write(ws, "AU11", year[0]); _safe_write(ws, "AW11", year[1])
        _safe_write(ws, "AY11", year[2]); _safe_write(ws, "BA11", year[3])
    tax_office = str(data.get("tax_office", ""))
    if len(tax_office) >= 4:
        _safe_write(ws, "BU11", tax_office[0]); _safe_write(ws, "BW11", tax_office[1])
        _safe_write(ws, "BY11", tax_office[2]); _safe_write(ws, "CA11", tax_office[3])
    _safe_write(ws, "K16", "6"); _safe_write(ws, "M16", "4"); _safe_write(ws, "O16", "3")
    _safe_write(ws, "AU16", "7"); _safe_write(ws, "AW16", "6"); _safe_write(ws, "AY16", "0")
    _write_fio_field(ws, data.get("last_name", ""), 11, 18)
    _write_fio_field(ws, data.get("first_name", ""), 11, 20)
    _write_fio_field(ws, data.get("middle_name", ""), 11, 22)
    birth_date = str(data.get("birth_date", ""))
    if len(birth_date) == 10:
        _safe_write(ws, "K31", birth_date[0]); _safe_write(ws, "M31", birth_date[1])
        _safe_write(ws, "Q31", birth_date[3]); _safe_write(ws, "S31", birth_date[4])
        _safe_write(ws, "W31", birth_date[6]); _safe_write(ws, "Y31", birth_date[7])
        _safe_write(ws, "AA31", birth_date[8]); _safe_write(ws, "AC31", birth_date[9])
    _safe_write(ws, "Q33", "2"); _safe_write(ws, "S33", "1")
    passport = str(data.get("passport", ""))
    if len(passport) == 10:
        col = 17
        for digit in passport:
            _safe_write(ws, f"{get_column_letter(col)}35", digit)
            col += 2
    _safe_write(ws, "Y37", "1")
    _write_fio_field(ws, data.get("taxpayer_phone", ""), 21, 39)
    pages_str = str(total_pages).zfill(3)
    _safe_write(ws, "S44", pages_str[0]); _safe_write(ws, "U44", pages_str[1]); _safe_write(ws, "W44", pages_str[2])
    _safe_write(ws, "BQ44", "?"); _safe_write(ws, "BS44", "?"); _safe_write(ws, "BU44", "?")
    ws["BQ44"].fill = GREEN_FILL; ws["BS44"].fill = GREEN_FILL; ws["BU44"].fill = GREEN_FILL
    _safe_write(ws, "D48", "1")
    _write_fio_field(ws, data.get("last_name", ""), 2, 50)
    _write_fio_field(ws, data.get("first_name", ""), 2, 52)
    _write_fio_field(ws, data.get("middle_name", ""), 2, 54)
    today = datetime.now().strftime("%d%m%Y")
    _safe_write(ws, "V56", today[0]); _safe_write(ws, "X56", today[1])
    _safe_write(ws, "AB56", today[2]); _safe_write(ws, "AD56", today[3])
    _safe_write(ws, "AH56", today[4]); _safe_write(ws, "AJ56", today[5])
    _safe_write(ws, "AL56", today[6]); _safe_write(ws, "AN56", today[7])


def _write_inn(ws, inn):
    inn = str(inn) if inn else ""
    if len(inn) < 12: return
    cols = [25, 27, 29, 31, 33, 35, 37, 39, 41, 43, 45, 47]
    for i, digit in enumerate(inn[:12]):
        _safe_write(ws, f"{get_column_letter(cols[i])}1", digit)


# ==================== РАЗДЕЛ 1 ====================

def _fill_section1(wb, data):
    ws = wb["Раздел 1"]
    _write_fio_section_header(ws, data)
    _write_number_field(ws, "18210102010011000110", 21, 12)
    tax_to_pay = data.get("tax_to_pay", 0); tax_return = data.get("tax_return", 0)
    if tax_to_pay > 0: _write_number_field(ws, str(round(tax_to_pay)), 21, 16)
    elif tax_return > 0: _write_number_field(ws, str(round(tax_return)), 21, 18)
    _safe_write(ws, "V63", datetime.now().strftime("%d.%m.%Y"), font_size=8)


def _fill_return_request(wb, data):
    ws = wb["Прил-е к Разделу 1"]
    _write_fio_section_header(ws, data)
    _write_number_field(ws, str(round(data.get("tax_return", 0))), 13, 11)
    bik = str(data.get("bik", ""))
    if len(bik) == 9: _write_number_field(ws, bik, 21, 15)
    account = str(data.get("account", ""))
    if len(account) == 20: _write_number_field(ws, account, 21, 17)
    card = str(data.get("card", ""))
    if card: _write_number_field(ws, card, 21, 19)
    _safe_write(ws, "V50", datetime.now().strftime("%d.%m.%Y"), font_size=8)


def _fill_section2(wb, data):
    ws = wb["Раздел 2"]
    _write_fio_section_header(ws, data)
    _safe_write(ws, "Y9", "0"); _safe_write(ws, "Z9", "1")
    income = data.get("income", 0)
    _write_amount_with_kopeks(ws, income, 25, 11)
    _write_amount_with_kopeks(ws, data.get("total_deduction", 0), 25, 17)
    tax_base = max(0, income - data.get("total_deduction", 0))
    _write_amount_with_kopeks(ws, tax_base, 25, 21)
    _write_number_field(ws, str(round(tax_base * 0.13)), 25, 24)
    _write_number_field(ws, str(round(data.get("tax_paid", 0))), 25, 26)
    tax_to_pay = data.get("tax_to_pay", 0); tax_return = data.get("tax_return", 0)
    if tax_to_pay > 0: _write_number_field(ws, str(round(tax_to_pay)), 25, 38)
    if tax_return > 0: _write_number_field(ws, str(round(tax_return)), 25, 40)
    _safe_write(ws, "V59", datetime.now().strftime("%d.%m.%Y"), font_size=8)


def _fill_appendix5_education(wb, data):
    ws = wb["Прил.5"]
    _write_fio_section_header(ws, data)
    _write_amount_with_kopeks(ws, data.get("education_total", 0), 26, 42)
    _safe_write(ws, "V47", datetime.now().strftime("%d.%m.%Y"), font_size=8)


def _fill_appendix5_medical(wb, data):
    ws = wb["Прил.5 (продолжение)"]
    _write_fio_section_header(ws, data)
    mt = data.get("medical_total", 0)
    _write_amount_with_kopeks(ws, mt, 26, 9)
    _write_amount_with_kopeks(ws, mt, 26, 23)
    _write_amount_with_kopeks(ws, mt, 26, 29)
    _write_amount_with_kopeks(ws, mt, 26, 31)
    _safe_write(ws, "V54", datetime.now().strftime("%d.%m.%Y"), font_size=8)


def _fill_appendix5_investment(wb, data):
    ws = wb["Прил.5"]
    _write_fio_section_header(ws, data)
    _write_amount_with_kopeks(ws, data.get("investment_amount", 0), 26, 42)
    _safe_write(ws, "V47", datetime.now().strftime("%d.%m.%Y"), font_size=8)


def _fill_calc_appendix5(wb, data):
    ws = wb["Расчет к прил.5"]
    _write_fio_section_header(ws, data)

    inv_amount = data.get("investment_amount", 0)
    broker_inn = str(data.get("investment_broker_inn", ""))
    broker_name = str(data.get("investment_broker_name", ""))
    contract = str(data.get("investment_contract", ""))
    open_date = str(data.get("investment_open_date", ""))

    # ИНН брокера (строка 090)
    if len(broker_inn) >= 10:
        _write_number_field(ws, broker_inn, 1, 9)

    # Название брокера (строка 110)
    if broker_name:
        _safe_write(ws, "A11", broker_name)

    # Дата договора (строка 120)
    if open_date and len(open_date) == 10:
        _safe_write(ws, "A12", open_date)

    # Номер договора (строка 130)
    if contract:
        _safe_write(ws, "A13", contract)

    # Сумма взноса (строка 150)
    _write_amount_with_kopeks(ws, inv_amount, 26, 15)

    # Признак основания (строка 080) — 1 (статья 219.1)
    _safe_write(ws, "D8", "1")

    _safe_write(ws, "V47", datetime.now().strftime("%d.%m.%Y"), font_size=8)


def _fill_appendix7(wb, data):
    ws = wb["Прил.7"]
    _write_fio_section_header(ws, data)

    pp = data.get("property_price", data.get("property_total", 0))
    pm = data.get("property_mortgage", 0)
    pot = data.get("property_object_type", "5")
    pc = data.get("property_cadastral", "")
    pa = data.get("property_address", "")
    pad = data.get("property_act_date", "")
    prd = data.get("property_reg_date", "")
    income = data.get("income", 0)
    td = data.get("total_deduction", 0)

    _safe_write(ws, "K10", pot); _safe_write(ws, "Y10", "0"); _safe_write(ws, "Z10", "1")
    _safe_write(ws, "Y12", "2")
    if pc: _write_number_field(ws, pc, 1, 14)
    if pa:
        lines = [pa[i:i+40] for i in range(0, len(pa), 40)]
        for idx, line in enumerate(lines[:7]):
            row = 16 + idx * 2
            for col_idx, char in enumerate(line):
                if col_idx >= 40: break
                _safe_write(ws, f"{get_column_letter(col_idx + 1)}{row}", char)
    if pad and len(pad) == 10:
        _safe_write(ws, "AD30", pad[0]); _safe_write(ws, "AE30", pad[1])
        _safe_write(ws, "AG30", pad[3]); _safe_write(ws, "AH30", pad[4])
        _safe_write(ws, "AJ30", pad[6]); _safe_write(ws, "AK30", pad[7])
        _safe_write(ws, "AL30", pad[8]); _safe_write(ws, "AM30", pad[9])
    if prd and len(prd) == 10:
        _safe_write(ws, "AD32", prd[0]); _safe_write(ws, "AE32", prd[1])
        _safe_write(ws, "AG32", prd[3]); _safe_write(ws, "AH32", prd[4])
        _safe_write(ws, "AJ32", prd[6]); _safe_write(ws, "AK32", prd[7])
        _safe_write(ws, "AL32", prd[8]); _safe_write(ws, "AM32", prd[9])
    ded_price = min(pp, 2_000_000)
    _write_amount_with_kopeks(ws, ded_price, 30, 34)
    if pm > 0:
        _write_amount_with_kopeks(ws, min(pm, 3_000_000), 30, 36)
    _write_amount_with_kopeks(ws, max(0, income - td), 26, 51)
    _write_amount_with_kopeks(ws, ded_price, 30, 53)
    _write_amount_with_kopeks(ws, max(0, pp - ded_price), 30, 57)
    _safe_write(ws, "V63", datetime.now().strftime("%d.%m.%Y"), font_size=8)