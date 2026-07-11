import re
import os
import openpyxl
from bot.config import MEDICAL_KEYWORDS, EDUCATION_KEYWORDS, DATA_TEMP_DIR


async def parse_excel(file_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    debug_path = os.path.join(DATA_TEMP_DIR, "excel_debug.txt")
    debug_lines = []

    # Ищем заголовки и определяем колонки
    header_row, col_date, col_desc, col_amount = _find_headers(ws)
    debug_lines.append(f"Заголовки: строка={header_row}, дата={col_date}, описание={col_desc}, сумма={col_amount}")

    if header_row is None:
        _write_debug(debug_path, debug_lines)
        wb.close()
        return []

    payments = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or all(cell is None for cell in row):
            continue

        cells = [str(c).strip() if c is not None else "" for c in row]

        # Дата
        date = None
        if col_date is not None and col_date < len(cells):
            date = _parse_date(cells[col_date])

        # Описание
        description = ""
        if col_desc is not None and col_desc < len(cells):
            description = cells[col_desc]
            description = description.replace('"', "").strip()
            description = re.sub(r"\s+", " ", description)

        # Сумма
        amount = None
        if col_amount is not None and col_amount < len(cells):
            amount = _parse_amount(cells[col_amount])

        if not date or amount is None or not description:
            continue

        category = _detect_category(description)

        debug_lines.append(f"OK: {date} | {amount:>12.2f} | {description[:100]} | {category}")

        payments.append({
            "date": date,
            "amount": abs(amount) if amount < 0 else amount,
            "description": description,
            "category": category
        })

    debug_lines.append(f"Всего платежей: {len(payments)}")
    _write_debug(debug_path, debug_lines)

    wb.close()
    return payments


def _write_debug(path, lines):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def _find_headers(ws) -> tuple:
    """
    Ищет строку заголовков и возвращает (row_index, col_date, col_desc, col_amount).
    Ищет по ключевым словам в первых 20 строках.
    """
    date_keywords = ["дата", "date", "день", "day"]
    desc_keywords = ["назначение", "описание", "description", "платеж", "получатель", "отправитель", "payee", "narrative"]
    amount_keywords = ["сумма", "amount", "рубл", "rub", "₽", "сумма в валюте", "российские рубли"]

    for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
        if not row:
            continue
        cells = [str(c).lower().strip() if c is not None else "" for c in row]

        col_date = None
        col_desc = None
        col_amount = None

        for i, cell in enumerate(cells):
            # Ищем дату
            if col_date is None and any(kw in cell for kw in date_keywords):
                col_date = i
            # Ищем описание
            if col_desc is None and any(kw in cell for kw in desc_keywords):
                col_desc = i
            # Ищем сумму (приоритет — российские рубли)
            if any(kw in cell for kw in ["российские рубли", "рубл"]):
                col_amount = i
            elif col_amount is None and any(kw in cell for kw in amount_keywords):
                col_amount = i

        # Если нашли хотя бы дату и сумму — это заголовок
        if col_date is not None and col_amount is not None:
            return row_idx, col_date, col_desc, col_amount

    return None, None, None, None


def _parse_date(text: str) -> str | None:
    """Парсит дату из строки в любом формате, возвращает ДД.ММ.ГГГГ."""
    if not text:
        return None
    text = text.strip()

    # ISO: 2025-03-18, 2025-03-18 14:24:00
    match = re.search(r"(\d{4})-(\d{2})-(\d{2})", text)
    if match:
        return f"{match.group(3)}.{match.group(2)}.{match.group(1)}"

    # Российский: 18.03.2025, 18.03.2025 14:24:00
    match = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", text)
    if match:
        return f"{match.group(1)}.{match.group(2)}.{match.group(3)}"

    # Американский: 03/18/2025
    match = re.search(r"(\d{2})/(\d{2})/(\d{4})", text)
    if match:
        return f"{match.group(1)}.{match.group(2)}.{match.group(3)}"

    return None


def _parse_amount(text: str) -> float | None:
    """
    Извлекает сумму из ячейки.
    Поддерживает форматы:
    - 225.00 ₽, + 1 500.00 ₽, -200 000,00 ₽
    -225.00, 1500.00-, 200000
    """
    if not text:
        return None
    text = text.strip()

    # Вариант 1: со знаком ₽
    match = re.search(r"([+-]?)\s*(\d{1,3}(?:\s?\d{3})*(?:[.,]\d{2})?)\s*₽", text)
    if match:
        sign = match.group(1) or "+"
        raw = match.group(2).replace(" ", "").replace(",", ".")
        try:
            amount = float(raw)
            return -amount if sign == "-" else amount
        except ValueError:
            pass

    # Вариант 2: число с минусом в конце (1500.00-)
    match = re.search(r"(\d{1,3}(?:\s?\d{3})*(?:[.,]\d{2})?)\s*-", text)
    if match:
        raw = match.group(1).replace(" ", "").replace(",", ".")
        try:
            return -float(raw)
        except ValueError:
            pass

    # Вариант 3: просто число со знаком
    match = re.search(r"([+-])\s*(\d{1,3}(?:\s?\d{3})*(?:[.,]\d{2})?)", text)
    if match:
        raw = match.group(2).replace(" ", "").replace(",", ".")
        try:
            amount = float(raw)
            return -amount if match.group(1) == "-" else amount
        except ValueError:
            pass

    return None


def _detect_category(description: str) -> str | None:
    desc_lower = description.lower()

    for kw in MEDICAL_KEYWORDS:
        if kw.lower() in desc_lower:
            return "medical"

    for kw in EDUCATION_KEYWORDS:
        if kw.lower() in desc_lower:
            return "education"

    medical_patterns = [
        r"гбуз", r"г\s*б\s*у\s*з", r"поликлин", r"госпитал",
        r"диспансер", r"роддом", r"мед\s*центр", r"стоматолог", r"зубн",
        r"тгкб", r"гкб", r"црб", r"ркб", r"клиник", r"больниц",
        r"медицин", r"аптек", r"лечени", r"диагност", r"анализ",
        r"хирург", r"терапевт", r"врач", r"медосмотр", r"санатор"
    ]
    for pattern in medical_patterns:
        if re.search(pattern, desc_lower):
            return "medical"

    education_patterns = [
        r"универ", r"институт", r"академи", r"колледж",
        r"школ", r"гимназ", r"лицей", r"вуз", r"образован",
        r"обучен", r"курс", r"тренинг", r"семинар", r"репетитор",
        r"автошкол", r"музыкальн", r"спортивн", r"техникум", r"училищ"
    ]
    for pattern in education_patterns:
        if re.search(pattern, desc_lower):
            return "education"

    return None